#!/usr/bin/env python3
"""Parse Kayls delivery + tracker workbooks for Shopify import."""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

KAYLS = Path(r"C:\Users\ghadi.mdallal\Downloads\Kayls Delivery  Orders (1).xlsx")
TRACKER = Path(r"C:\Users\ghadi.mdallal\Downloads\Scent_Square_Tracker_REPAIRED.xlsx")
OUT = Path(r"C:\myProjects\supplify_erp\tmp_kayls_parsed.json")


def sheet_to_rows(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([("" if c is None else c) for c in row])
        result[name] = rows
    return result


def main():
    downloads = Path(r"C:\Users\ghadi.mdallal\Downloads")
    matches = {
        "kayls_matches": [str(p) for p in downloads.glob("*Kayls*")],
        "scent_matches": [str(p) for p in downloads.glob("*Scent*")],
        "tracker_matches": [str(p) for p in downloads.glob("*Tracker*")],
        "delivery_matches": [str(p) for p in downloads.glob("*Delivery*")],
    }

    kayls_path = KAYLS if KAYLS.exists() else None
    if not kayls_path:
        for p in downloads.glob("*Kayls*"):
            if p.suffix.lower() == ".xlsx":
                kayls_path = p
                break

    tracker_path = TRACKER if TRACKER.exists() else None
    if not tracker_path:
        for p in downloads.glob("*Tracker*"):
            if p.suffix.lower() == ".xlsx":
                tracker_path = p
                break

    payload = {
        "matches": matches,
        "kayls_path": str(kayls_path) if kayls_path else None,
        "tracker_path": str(tracker_path) if tracker_path else None,
        "kayls": sheet_to_rows(kayls_path) if kayls_path else None,
        "tracker": sheet_to_rows(tracker_path) if tracker_path else None,
    }
    OUT.write_text(json.dumps(payload, default=str, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")
    print(json.dumps({"kayls_path": payload["kayls_path"], "tracker_path": payload["tracker_path"], "matches": matches}, indent=2))


if __name__ == "__main__":
    main()
