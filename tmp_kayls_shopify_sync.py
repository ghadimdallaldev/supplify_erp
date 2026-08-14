#!/usr/bin/env python3
"""
Kayls delivery orders + sold-stock sync for Scent Square (e3y11v-zm.myshopify.com).

Implements the approved plan:
- One paid+fulfilled Shopify order per Kayls row (name/phone/total)
- Custom line titled from sold list + discount to match Kayls total
- Inventory 0 on confirmed sold products only (not BTSO/Bohoboco)

Run:
  python C:\\myProjects\\supplify_erp\\tmp_kayls_shopify_sync.py
"""
from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import tempfile
import time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

STORE = "e3y11v-zm.myshopify.com"
SHOPIFY = r"C:\Users\ghadi.mdallal\AppData\Roaming\npm\shopify.cmd"
AGENT_INFO = "n:cursor|v:none|p:none|m:composer"
AGENT_IDS = "s:1fc4c7e3-2428-4d17-a113-553244841c55|r:kayls-import|i:1"
WORK = Path(r"C:\myProjects\supplify_erp\tmp_kayls_work")
REPORT = WORK / "report.json"

KAYLS_CANDIDATES = [
    Path(r"C:\Users\ghadi.mdallal\Downloads\Kayls Delivery  Orders (1).xlsx"),
    Path(r"C:\Users\ghadi.mdallal\Downloads\Kayls Delivery Orders (1).xlsx"),
]
TRACKER_CANDIDATES = [
    Path(r"C:\Users\ghadi.mdallal\Downloads\Scent_Square_Tracker_REPAIRED.xlsx"),
]

# Confirmed sold list (canonical) — order line titles + inventory zero
SOLD_LIST = [
    "Cedrat – Real Tester",
    "Essential Parfums Velvet – Sealed",
    "Essential Parfums The Musc – Sealed",
    "Nishane Hacivat – 50 ml – Sealed",
    "Moschino Toy Boy – Sealed",
    "Dior Sauvage EDT – 100 ml – No Box – Refillable",
    "Hawas Éclat",
    "Kenzo Indigo",
    "Chopard Rose Malaki – Sealed",
    "Chopard Musk Malaki – Sealed",
    "YSL Libre EDP Intense – 200 ml – Open Box – With Box",
    "Gucci Intense Oud – With Box",
    "Chanel Coco Mademoiselle EDT – Real Tester",
    "YSL Libre Platine – Open Box",
    "Narciso Musc Noir Rose – 100 ml – No Box",
    "YSL MYSLF EDP",
    "Armani Sì EDP Intense – 2025 Edition",
    "Azzaro The Most Wanted EDP Intense – Open Box",
    "Bleu de Chanel Parfum – 150 ml",
    "Miss Dior EDP",
    "Lancôme La Vie Est Belle – 100 ml",
    "Prada L’Homme Intense EDP",
    "YSL Black Opium EDP",
    "YSL Black Opium EDT Intense",
    "Jean Paul Gaultier La Belle Le Parfum",
    "Prada L’Homme EDP",
    "My Burberry – 100 ml – Open Box",
    "Twilly d’Hermès",
    "Armani Sì EDP Intense",
    "Guerlain Homme EDP",
    "Bleu de Chanel EDP",
]

# Search keywords for Shopify product matching (avoid BTSO/Bohoboco)
SOLD_SEARCH = [
    ("Cedrat – Real Tester", "Cedrat"),
    ("Essential Parfums Velvet – Sealed", "Essential Velvet"),
    ("Essential Parfums The Musc – Sealed", "Essential Musc"),
    ("Nishane Hacivat – 50 ml – Sealed", "Hacivat"),
    ("Moschino Toy Boy – Sealed", "Toy Boy"),
    ("Dior Sauvage EDT – 100 ml – No Box – Refillable", "Sauvage EDT"),
    ("Hawas Éclat", "Hawas"),
    ("Kenzo Indigo", "Kenzo Indigo"),
    ("Chopard Rose Malaki – Sealed", "Rose Malaki"),
    ("Chopard Musk Malaki – Sealed", "Musk Malaki"),
    ("YSL Libre EDP Intense – 200 ml – Open Box – With Box", "Libre Intense"),
    ("Gucci Intense Oud – With Box", "Intense Oud"),
    ("Chanel Coco Mademoiselle EDT – Real Tester", "Coco Mademoiselle"),
    ("YSL Libre Platine – Open Box", "Libre Platine"),
    ("Narciso Musc Noir Rose – 100 ml – No Box", "Musc Noir Rose"),
    ("YSL MYSLF EDP", "MYSLF"),
    ("Armani Sì EDP Intense – 2025 Edition", "Si Intense 2025"),
    ("Azzaro The Most Wanted EDP Intense – Open Box", "Most Wanted"),
    ("Bleu de Chanel Parfum – 150 ml", "Bleu de Chanel Parfum"),
    ("Miss Dior EDP", "Miss Dior"),
    ("Lancôme La Vie Est Belle – 100 ml", "La Vie Est Belle"),
    ("Prada L’Homme Intense EDP", "L'Homme Intense"),
    ("YSL Black Opium EDP", "Black Opium EDP"),
    ("YSL Black Opium EDT Intense", "Black Opium EDT"),
    ("Jean Paul Gaultier La Belle Le Parfum", "La Belle Le Parfum"),
    ("Prada L’Homme EDP", "L'Homme EDP"),
    ("My Burberry – 100 ml – Open Box", "My Burberry"),
    ("Twilly d’Hermès", "Twilly"),
    ("Armani Sì EDP Intense", "Si EDP Intense"),
    ("Guerlain Homme EDP", "Guerlain Homme"),
    ("Bleu de Chanel EDP", "Bleu de Chanel EDP"),
]

SCOPES = ",".join(
    [
        "write_draft_orders",
        "read_draft_orders",
        "write_customers",
        "read_customers",
        "read_orders",
        "write_merchant_managed_fulfillment_orders",
        "read_merchant_managed_fulfillment_orders",
        "read_products",
        "read_inventory",
        "write_inventory",
    ]
)


def find_file(candidates: list[Path], glob_pat: str) -> Path:
    for c in candidates:
        if c.exists():
            return c
    downloads = Path(r"C:\Users\ghadi.mdallal\Downloads")
    hits = sorted(downloads.glob(glob_pat), key=lambda p: p.stat().st_mtime, reverse=True)
    if not hits:
        raise FileNotFoundError(f"No file matching {glob_pat} in Downloads")
    return hits[0]


def shopify_env() -> dict:
    env = os.environ.copy()
    env["SHOPIFY_CLI_AGENT_INFO"] = AGENT_INFO
    env["SHOPIFY_CLI_AGENT_IDS"] = AGENT_IDS
    return env


def shopify_execute(query: str, variables: dict | None = None, mutate: bool = False) -> dict:
    WORK.mkdir(parents=True, exist_ok=True)
    qpath = WORK / f"q_{int(time.time() * 1000)}.graphql"
    qpath.write_text(query, encoding="utf-8")
    cmd = [
        SHOPIFY if Path(SHOPIFY).exists() else "shopify.cmd",
        "store",
        "execute",
        "--store",
        STORE,
        "--query-file",
        str(qpath),
        "--json",
    ]
    vpath = None
    if variables is not None:
        vpath = WORK / f"v_{int(time.time() * 1000)}.json"
        vpath.write_text(json.dumps(variables), encoding="utf-8")
        cmd.extend(["--variable-file", str(vpath)])
    if mutate:
        cmd.append("--allow-mutations")
    r = subprocess.run(cmd, capture_output=True, text=True, env=shopify_env())
    out = (r.stdout or "") + ("\n" + r.stderr if r.stderr else "")
    # CLI may wrap JSON; find last JSON object
    data = None
    for line in out.splitlines()[::-1]:
        line = line.strip()
        if line.startswith("{") and line.endswith("}"):
            try:
                data = json.loads(line)
                break
            except json.JSONDecodeError:
                continue
    if data is None:
        # try whole stdout
        try:
            data = json.loads(r.stdout)
        except Exception:
            raise RuntimeError(f"shopify execute failed (code={r.returncode}):\n{out[:4000]}")
    if r.returncode != 0 and "data" not in data:
        raise RuntimeError(f"shopify execute failed (code={r.returncode}):\n{out[:4000]}")
    return data


def auth():
    cmd = [
        SHOPIFY if Path(SHOPIFY).exists() else "shopify.cmd",
        "store",
        "auth",
        "--store",
        STORE,
        "--scopes",
        SCOPES,
    ]
    r = subprocess.run(cmd, capture_output=True, text=True, env=shopify_env())
    print(r.stdout)
    if r.stderr:
        print(r.stderr)
    if r.returncode != 0:
        raise RuntimeError(f"shopify store auth failed: {r.returncode}")


def normalize_phone(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() == "nan":
        return None
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None
    if digits.startswith("961"):
        return f"+{digits}"
    if digits.startswith("0") and len(digits) >= 8:
        return f"+961{digits[1:]}"
    if len(digits) == 8:
        return f"+961{digits}"
    if s.startswith("+"):
        return f"+{digits}"
    return f"+{digits}"


def parse_kayls(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    print("Kayls headers:", headers)

    def col(*names):
        for n in names:
            for i, h in enumerate(headers):
                if n in h:
                    return i
        return None

    i_name = col("name", "customer", "client", "اسم")
    i_phone = col("phone", "mobile", "tel", "number", "رقم")
    i_total = col("total", "amount", "price", "cod", "value", "مبلغ")
    i_addr = col("address", "location", "area", "city", "عنوان")
    i_note = col("note", "notes", "remark")
    i_date = col("date")
    i_order = col("order", "ref", "id", "#")

    orders = []
    for ridx, row in enumerate(rows[1:], start=2):
        vals = list(row)
        def get(i):
            if i is None or i >= len(vals):
                return None
            return vals[i]

        name = get(i_name)
        phone = get(i_phone)
        total = get(i_total)
        if name is None and phone is None and total is None:
            continue
        # If name column missing, try first string cell
        if name is None:
            for v in vals:
                if isinstance(v, str) and v.strip():
                    name = v.strip()
                    break
        try:
            amount = float(str(total).replace("$", "").replace(",", "").strip()) if total is not None else 0.0
        except ValueError:
            amount = 0.0
        if amount <= 0:
            continue
        orders.append(
            {
                "row": ridx,
                "name": str(name).strip() if name else "Customer",
                "phone": normalize_phone(phone),
                "phone_raw": str(phone).strip() if phone is not None else None,
                "total": round(amount, 2),
                "address": str(get(i_addr)).strip() if get(i_addr) else None,
                "note": str(get(i_note)).strip() if get(i_note) else None,
                "date": str(get(i_date)) if get(i_date) else None,
                "order_ref": str(get(i_order)).strip() if get(i_order) else None,
                "raw": ["" if v is None else v for v in vals],
            }
        )
    return orders


def split_name(full: str) -> tuple[str, str]:
    parts = full.strip().split()
    if not parts:
        return ("Kayls", "Customer")
    if len(parts) == 1:
        return (parts[0], ".")
    return (parts[0], " ".join(parts[1:]))


def find_or_create_customer(name: str, phone: str | None, address: str | None) -> dict:
    first, last = split_name(name)
    nodes = []
    if phone:
        q = """
        query($q: String!) {
          customers(first: 10, query: $q) {
            nodes { id displayName phone defaultPhoneNumber { phoneNumber } defaultEmailAddress { emailAddress } }
          }
        }
        """
        # try several phone query forms
        variants = [phone, phone.replace("+", ""), phone[-8:]]
        for v in variants:
            data = shopify_execute(q, {"q": f"phone:{v}"})
            nodes = (((data.get("data") or {}).get("customers") or {}).get("nodes")) or []
            if nodes:
                break
    if not nodes:
        # name search
        q = """
        query($q: String!) {
          customers(first: 5, query: $q) {
            nodes { id displayName phone defaultPhoneNumber { phoneNumber } }
          }
        }
        """
        data = shopify_execute(q, {"q": name})
        nodes = (((data.get("data") or {}).get("customers") or {}).get("nodes")) or []

    if nodes:
        c = nodes[0]
        return {"id": c["id"], "displayName": c.get("displayName"), "created": False}

    inp = {"firstName": first, "lastName": last}
    if phone:
        inp["phone"] = phone
    if address:
        inp["addresses"] = [
            {
                "address1": address[:255],
                "countryCode": "LB",
                "firstName": first,
                "lastName": last,
                **({"phone": phone} if phone else {}),
            }
        ]
    mut = """
    mutation($input: CustomerInput!) {
      customerCreate(input: $input) {
        customer { id displayName }
        userErrors { field message }
      }
    }
    """
    data = shopify_execute(mut, {"input": inp}, mutate=True)
    payload = ((data.get("data") or {}).get("customerCreate")) or {}
    errs = payload.get("userErrors") or []
    if errs:
        # phone conflict: re-search
        msg = " ".join(e.get("message", "") for e in errs)
        if phone and "phone" in msg.lower():
            data2 = shopify_execute(
                """
                query($q: String!) {
                  customers(first: 5, query: $q) {
                    nodes { id displayName }
                  }
                }
                """,
                {"q": f"phone:{phone}"},
            )
            nodes2 = (((data2.get("data") or {}).get("customers") or {}).get("nodes")) or []
            if nodes2:
                return {"id": nodes2[0]["id"], "displayName": nodes2[0].get("displayName"), "created": False}
        raise RuntimeError(f"customerCreate errors for {name}: {errs}")
    cust = payload.get("customer")
    return {"id": cust["id"], "displayName": cust.get("displayName"), "created": True}


def search_products(term: str) -> list[dict]:
    q = """
    query($q: String!) {
      products(first: 15, query: $q, sortKey: CREATED_AT, reverse: true) {
        nodes {
          id title status createdAt
          variants(first: 3) {
            nodes {
              id title price inventoryQuantity
              inventoryItem { id }
            }
          }
        }
      }
    }
    """
    data = shopify_execute(q, {"q": f"title:{term}"})
    return (((data.get("data") or {}).get("products") or {}).get("nodes")) or []


def best_match(label: str, term: str) -> dict | None:
    nodes = search_products(term)
    # filter out BTSO / Bohoboco / new unsold brands
    blocked = ("born to stand", "bohoboco", "btso", "amouage", "layton", "omanluxury", "oman luxury")
    filtered = [n for n in nodes if not any(b in (n.get("title") or "").lower() for b in blocked)]
    if not filtered:
        # broader search without title:
        nodes = search_products(term.split()[0] if term else label)
        filtered = [n for n in nodes if not any(b in (n.get("title") or "").lower() for b in blocked)]
    if not filtered:
        return None
    # newest first already; prefer status ACTIVE with stock, else any
    filtered.sort(key=lambda n: n.get("createdAt") or "", reverse=True)
    return filtered[0]


def create_order(customer_id: str, phone: str | None, title: str, total: float, kayls_meta: dict) -> dict:
    # custom line at total, no discount needed if exact; still use discount=0 path via appliedDiscount only if needed
    line_price = f"{total:.2f}"
    inp = {
        "customerId": customer_id,
        "taxExempt": True,
        "note": f"Kayls delivery import | row={kayls_meta.get('row')} | ref={kayls_meta.get('order_ref')} | {kayls_meta.get('note') or ''}".strip(),
        "tags": ["Kayls", "imported-2026-08-14"],
        "lineItems": [
            {
                "title": title,
                "quantity": 1,
                "originalUnitPrice": line_price,
                "requiresShipping": True,
            }
        ],
    }
    if phone:
        inp["phone"] = phone
    if kayls_meta.get("address"):
        first, last = split_name(kayls_meta.get("name") or "Customer")
        addr = {
            "address1": kayls_meta["address"][:255],
            "countryCode": "LB",
            "firstName": first,
            "lastName": last,
        }
        if phone:
            addr["phone"] = phone
        inp["shippingAddress"] = addr
        inp["billingAddress"] = addr

    mut = """
    mutation($input: DraftOrderInput!) {
      draftOrderCreate(input: $input) {
        draftOrder { id name totalPrice }
        userErrors { field message }
      }
    }
    """
    data = shopify_execute(mut, {"input": inp}, mutate=True)
    payload = ((data.get("data") or {}).get("draftOrderCreate")) or {}
    errs = payload.get("userErrors") or []
    if errs:
        raise RuntimeError(f"draftOrderCreate: {errs}")
    draft = payload["draftOrder"]

    mut2 = """
    mutation($id: ID!, $paymentPending: Boolean) {
      draftOrderComplete(id: $id, paymentPending: $paymentPending) {
        draftOrder { id order { id name } }
        userErrors { field message }
      }
    }
    """
    data2 = shopify_execute(mut2, {"id": draft["id"], "paymentPending": False}, mutate=True)
    payload2 = ((data2.get("data") or {}).get("draftOrderComplete")) or {}
    errs2 = payload2.get("userErrors") or []
    if errs2:
        raise RuntimeError(f"draftOrderComplete: {errs2}")
    order = (payload2.get("draftOrder") or {}).get("order")
    if not order:
        raise RuntimeError(f"No order from draft complete: {data2}")

    # fulfill
    qfo = """
    query($id: ID!) {
      order(id: $id) {
        id name
        fulfillmentOrders(first: 5) {
          nodes { id status }
        }
      }
    }
    """
    fo_data = shopify_execute(qfo, {"id": order["id"]})
    fo_nodes = (
        (((fo_data.get("data") or {}).get("order") or {}).get("fulfillmentOrders") or {}).get("nodes")
    ) or []
    open_fos = [f for f in fo_nodes if (f.get("status") or "").upper() in ("OPEN", "IN_PROGRESS", "SCHEDULED")]
    if not open_fos and fo_nodes:
        open_fos = fo_nodes
    fulfilled = False
    if open_fos:
        mut3 = """
        mutation($fulfillment: FulfillmentInput!) {
          fulfillmentCreate(fulfillment: $fulfillment) {
            fulfillment { id status }
            userErrors { field message }
          }
        }
        """
        f_in = {
            "lineItemsByFulfillmentOrder": [{"fulfillmentOrderId": open_fos[0]["id"]}],
            "notifyCustomer": False,
        }
        f_data = shopify_execute(mut3, {"fulfillment": f_in}, mutate=True)
        f_payload = ((f_data.get("data") or {}).get("fulfillmentCreate")) or {}
        f_errs = f_payload.get("userErrors") or []
        if f_errs:
            print(f"  warn fulfillment: {f_errs}")
        else:
            fulfilled = True

    return {
        "draft_id": draft["id"],
        "draft_name": draft.get("name"),
        "order_id": order["id"],
        "order_name": order.get("name"),
        "line_title": title,
        "total": total,
        "fulfilled": fulfilled,
    }


def zero_inventory(product: dict, location_id: str) -> dict:
    variants = ((product.get("variants") or {}).get("nodes")) or []
    results = []
    for v in variants:
        inv_item = (v.get("inventoryItem") or {}).get("id")
        qty = v.get("inventoryQuantity")
        if not inv_item:
            continue
        if qty is not None and int(qty) == 0:
            results.append({"variant": v.get("id"), "already_zero": True})
            continue
        mut = """
        mutation($input: InventorySetQuantitiesInput!) {
          inventorySetQuantities(input: $input) {
            inventoryAdjustmentGroup { reason }
            userErrors { field message }
          }
        }
        """
        variables = {
            "input": {
                "name": "available",
                "reason": "correction",
                "ignoreCompareQuantity": True,
                "quantities": [
                    {
                        "inventoryItemId": inv_item,
                        "locationId": location_id,
                        "quantity": 0,
                    }
                ],
            }
        }
        data = shopify_execute(mut, variables, mutate=True)
        payload = ((data.get("data") or {}).get("inventorySetQuantities")) or {}
        errs = payload.get("userErrors") or []
        results.append({"variant": v.get("id"), "errors": errs, "was_qty": qty})
    return {"product_id": product.get("id"), "title": product.get("title"), "results": results}


def get_location_id() -> str:
    data = shopify_execute(
        """
        query {
          locations(first: 10) {
            nodes { id name isActive }
          }
        }
        """
    )
    nodes = (((data.get("data") or {}).get("locations") or {}).get("nodes")) or []
    if not nodes:
        raise RuntimeError("No locations found")
    for n in nodes:
        if n.get("isActive"):
            return n["id"]
    return nodes[0]["id"]


def main():
    WORK.mkdir(parents=True, exist_ok=True)
    kayls_path = find_file(KAYLS_CANDIDATES, "*Kayls*.xlsx")
    print("Kayls file:", kayls_path)
    try:
        tracker_path = find_file(TRACKER_CANDIDATES, "*Tracker*.xlsx")
        print("Tracker file:", tracker_path)
    except FileNotFoundError:
        tracker_path = None
        print("Tracker file not found (optional)")

    orders = parse_kayls(kayls_path)
    print(f"Parsed {len(orders)} Kayls orders with total > 0")
    (WORK / "kayls_orders.json").write_text(json.dumps(orders, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    print("Auth...")
    auth()

    location_id = get_location_id()
    print("Location:", location_id)

    # Match sold products
    matched = []
    unmatched = []
    for label, term in SOLD_SEARCH:
        p = best_match(label, term)
        if p:
            matched.append({"label": label, "term": term, "product": p})
            print(f"MATCH {label} -> {p.get('title')} ({p.get('id')})")
        else:
            unmatched.append({"label": label, "term": term})
            print(f"NO MATCH {label}")

    (WORK / "sold_matches.json").write_text(
        json.dumps({"matched": matched, "unmatched": unmatched}, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    # Create orders
    line_titles = [m["product"]["title"] for m in matched] or SOLD_LIST
    created_orders = []
    customers_summary = []
    for i, o in enumerate(orders):
        title = line_titles[i % len(line_titles)]
        print(f"\nOrder {i+1}/{len(orders)}: {o['name']} {o['phone']} ${o['total']} -> {title}")
        cust = find_or_create_customer(o["name"], o["phone"], o.get("address"))
        customers_summary.append({**cust, "name": o["name"], "phone": o["phone"]})
        print(f"  customer {'CREATED' if cust['created'] else 'MATCHED'}: {cust['id']}")
        try:
            result = create_order(cust["id"], o["phone"], title, o["total"], o)
            created_orders.append({**result, "customer": cust, "kayls": o})
            print(f"  created {result['order_name']} fulfilled={result['fulfilled']}")
        except Exception as e:
            created_orders.append({"error": str(e), "kayls": o, "customer": cust, "line_title": title})
            print(f"  ERROR: {e}")
        time.sleep(0.3)

    # Zero inventory on matched sold products
    zeroed = []
    for m in matched:
        print(f"Zero stock: {m['product'].get('title')}")
        try:
            z = zero_inventory(m["product"], location_id)
            zeroed.append(z)
        except Exception as e:
            zeroed.append({"error": str(e), "label": m["label"], "product": m["product"].get("title")})
            print(f"  ERROR zero: {e}")
        time.sleep(0.2)

    report = {
        "kayls_path": str(kayls_path),
        "tracker_path": str(tracker_path) if tracker_path else None,
        "orders_count": len(orders),
        "created_orders": created_orders,
        "customers": customers_summary,
        "matched_sold": [{"label": m["label"], "title": m["product"].get("title"), "id": m["product"].get("id")} for m in matched],
        "unmatched_sold": unmatched,
        "zeroed": zeroed,
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print("\n=== DONE ===")
    print("Report:", REPORT)
    ok = [c for c in created_orders if "order_name" in c]
    err = [c for c in created_orders if "error" in c]
    print(f"Orders created: {len(ok)}  errors: {len(err)}")
    print(f"Customers created: {sum(1 for c in customers_summary if c.get('created'))}")
    print(f"Sold matched: {len(matched)} unmatched: {len(unmatched)}")


if __name__ == "__main__":
    main()
